# -*- coding: utf-8 -*-
# Copyright 2019 Ecosoft Co., Ltd (http://ecosoft.co.th)
# Copyright 2024 Jupetern
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl)

import logging

from odoo import api, models

_logger = logging.getLogger(__name__)

try:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.styles.borders import Border, Side
except ImportError:
    _logger.debug('Cannot import "openpyxl". Please make sure it is installed.')


class XLSXStyles(models.AbstractModel):
    _name = 'xlsx.styles'
    _description = 'Available styles for excel'

    @api.model
    def get_openpyxl_styles(self):
        """List all syles that can be used with styleing directive #{...}"""
        return {
            'font': {
                'arial_t': Font(name='Arial', size=16, color='000000', bold=True),
                'arial_h': Font(name='Arial', size=10, color='000000', bold=True),
                'arial_c': Font(name='Arial', size=10, color='000000', bold=False),
                'arial_8': Font(name='Arial', size=8, bold=False),
                'arial_8b': Font(name='Arial', size=8, bold=True),
                'arial_9': Font(name='Arial', size=9, bold=False),
                'arial_9b': Font(name='Arial', size=9, bold=True),
                'tnr_10': Font(name='Times New Roman', size=10, bold=False),
                'tnr_10b': Font(name='Times New Roman', size=10, bold=True),
            },
            'fill': {
                'grey_100': PatternFill('solid', fgColor='F0F0F0'),
                'red': PatternFill('solid', fgColor='FF0000'),
                'grey': PatternFill('solid', fgColor='DDDDDD'),
                'blue': PatternFill('solid', fgColor='9BF3FF'),
                'yellow': PatternFill('solid', fgColor='FFFCB7'),
                'green': PatternFill('solid', fgColor='B0FF99'),
            },
            'border': {
                'thin': Border(
                    left=Side(style='thin', color='808080'),
                    right=Side(style='thin', color='808080'),
                    top=Side(style='thin', color='808080'),
                    bottom=Side(style='thin', color='808080'),
                ),
            },
            'align': {
                'left': Alignment(horizontal='left', vertical='center'),
                'center': Alignment(horizontal='center', vertical='center'),
                'right': Alignment(horizontal='right', vertical='center'),
            },
            'format': {
                'qty': '#,##0.00',
                'number': '#,##0',
                'date': 'dd/mm/yyyy',
                'datestamp': 'yyyy-mm-dd',
                'text': '@',
                'percent': '0.00%',
            },
        }
